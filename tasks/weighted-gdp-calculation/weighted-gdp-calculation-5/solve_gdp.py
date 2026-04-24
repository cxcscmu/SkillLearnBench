#!/usr/bin/env python3
"""
Oracle solution for weighted GDP calculation.

Populates the Task sheet with computed values for:
- Step 1: Data lookup from the Data sheet (exports, imports by country/year)
- Step 2: Trade coefficient + statistics (MIN/MAX/MEDIAN/AVERAGE/PERCENTILE)
- Step 3: Weighted mean using SUMPRODUCT logic
"""

from openpyxl import load_workbook

EXCEL_FILE = "environment/gdp.xlsx"
EXCEL_OUT_FILE = "environment/gdp_answer.xlsx"

def main():
    wb = load_workbook(EXCEL_FILE)

    task_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Task' or 'Task' in sheet_name:
            task_sheet = wb[sheet_name]
            break
    if task_sheet is None:
        task_sheet = wb.active
    ws = task_sheet

    data_ws = None
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Data':
            data_ws = wb[sheet_name]
            break
    if data_ws is None:
        print("ERROR: Data sheet not found!")
        return

    columns = ['H', 'I', 'J', 'K', 'L', 'M']
    years = [2018, 2019, 2020, 2021, 2022, 2023]
    year_row = 9

    for col_idx, col in enumerate(columns):
        ws[f'{col}{year_row}'] = years[col_idx]

    # Map years to their column indices in the Data sheet
    year_to_col = {}
    for col_idx in range(1, 14):
        cell_val = data_ws.cell(row=4, column=col_idx).value
        if cell_val in years:
            year_to_col[cell_val] = col_idx

    # Map series codes to row numbers in the Data sheet (rows 21-33)
    series_to_row = {}
    for row in range(21, 34):
        series_code = data_ws.cell(row=row, column=2).value
        if series_code:
            series_to_row[series_code] = row

    def lookup_value(series_code, year):
        if series_code not in series_to_row:
            return 0
        if year not in year_to_col:
            return 0
        row = series_to_row[series_code]
        col = year_to_col[year]
        val = data_ws.cell(row=row, column=col).value
        return val if val is not None else 0

    export_rows = list(range(12, 18))
    import_rows = list(range(19, 25))

    # Step 1: Populate exports, imports values from Data sheet lookups
    for row in export_rows:
        series_code = ws.cell(row=row, column=4).value
        for col_idx, col in enumerate(columns):
            year = years[col_idx]
            value = lookup_value(series_code, year)
            ws[f'{col}{row}'] = value

    for row in import_rows:
        series_code = ws.cell(row=row, column=4).value
        for col_idx, col in enumerate(columns):
            year = years[col_idx]
            value = lookup_value(series_code, year)
            ws[f'{col}{row}'] = value


    # Step 2: Calculate trade coefficient = (exports - imports) / (exports + imports) * 100
    net_export_rows = list(range(28, 34))
    for row_idx, row in enumerate(net_export_rows):
        export_row = 12 + row_idx
        import_row = 19 + row_idx

        for col_idx, col in enumerate(columns):
            export_val = ws[f'{col}{export_row}'].value or 0
            import_val = ws[f'{col}{import_row}'].value or 0

            trade_coefficient = (export_val - import_val) / (export_val + import_val) * 100
            ws[f'{col}{row}'] = round(trade_coefficient, 1)

    # Step 2 continued: Compute statistics (MIN/MAX/MEDIAN/AVERAGE/PERCENTILE)
    for col_idx, col in enumerate(columns):
        values = [ws[f'{col}{r}'].value or 0 for r in range(28, 34)]
        ws[f'{col}35'] = min(values)
        ws[f'{col}36'] = max(values)
        sorted_vals = sorted(values)
        ws[f'{col}37'] = (sorted_vals[2] + sorted_vals[3]) / 2
        ws[f'{col}38'] = round(sum(values) / len(values), 1)
        ws[f'{col}39'] = sorted_vals[1] + 0.25 * (sorted_vals[2] - sorted_vals[1])
        ws[f'{col}40'] = sorted_vals[3] + 0.75 * (sorted_vals[4] - sorted_vals[3])

    # Step 3: GDP-weighted mean = SUMPRODUCT(trade coefficient, (exports + imports)) / SUM(exports + imports)
    for col_idx, col in enumerate(columns):
        trade_coefficients = [ws[f'{col}{r}'].value or 0 for r in range(28, 34)]
        trade_values = [ws[f'{col}{r}'].value + ws[f'{col}{r+7}'].value or 0 for r in range(12, 18)]

        sumproduct = sum(pct * trade for pct, trade in zip(trade_coefficients, trade_values))
        sum_trade = sum(trade_values)
        weighted_mean = sumproduct / sum_trade if sum_trade != 0 else 0
        ws[f'{col}43'] = round(weighted_mean, 1)

    wb.save(EXCEL_OUT_FILE)
    wb.close()
    print("Successfully computed all values.")

if __name__ == '__main__':
    main()
