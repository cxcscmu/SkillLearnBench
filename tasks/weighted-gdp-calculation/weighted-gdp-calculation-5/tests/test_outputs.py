"""
Tests for weighted GDP calculation task.
Verifies that the Excel formulas are correctly implemented for:
- Step 1: Lookup formulas (INDEX/MATCH, VLOOKUP/MATCH, etc.) for data retrieval
- Step 2: Net exports as % of GDP calculations and statistics
- Step 3: Weighted mean calculation using SUMPRODUCT
"""

import csv
import glob
from pathlib import Path

import pytest
from openpyxl import load_workbook

EXCEL_FILE = Path("/root/gdp.xlsx")
CSV_PATTERN = "/root/sheet.csv.*"
TOLERANCE = 0.5

_csv_data_cache = None
_task_sheet_index = None


def find_task_csv():
    """
    Locate the CSV file containing Task sheet data exported by ssconvert.

    Uses three strategies in order of reliability:
    1. Match by sheet index from xlsx (ssconvert names files sheet.csv.0, sheet.csv.1, etc.)
    2. Content pattern matching for Task sheet identifiers
    3. Fall back to first available CSV
    """
    global _task_sheet_index

    csv_files = sorted(glob.glob(CSV_PATTERN))
    if not csv_files:
        return None

    wb = load_workbook(EXCEL_FILE, data_only=False)
    for idx, name in enumerate(wb.sheetnames):
        if "Task" in name:
            _task_sheet_index = idx
            wb.close()
            expected_file = f"/root/sheet.csv.{idx}"
            if Path(expected_file).exists():
                return expected_file
            break
    wb.close()

    for csv_file in csv_files:
        try:
            with open(csv_file, encoding="utf-8", errors="ignore") as f:
                content = f.read(5000)
                if "466BXGS_BP6.A" in content or "United Arab Emirates" in content:
                    if "Exports, Goods and Services" in content:
                        return csv_file
        except:
            continue

    return csv_files[0] if csv_files else None


def load_csv_data():
    """
    Load and cache CSV data with evaluated formula values.

    Returns a dict mapping cell references (e.g., 'H12') to their values,
    converting numeric strings to floats where possible.
    """
    global _csv_data_cache

    if _csv_data_cache is not None:
        return _csv_data_cache

    csv_file = find_task_csv()
    if csv_file is None:
        _csv_data_cache = {}
        return _csv_data_cache

    _csv_data_cache = {}
    try:
        with open(csv_file, encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, val in enumerate(row):
                    col_letter = chr(ord("A") + col_idx) if col_idx < 26 else None
                    if col_letter:
                        cell_ref = f"{col_letter}{row_idx}"
                        if val and val.strip():
                            try:
                                _csv_data_cache[cell_ref] = float(val)
                            except ValueError:
                                _csv_data_cache[cell_ref] = val
                        else:
                            _csv_data_cache[cell_ref] = None
    except Exception as e:
        print(f"Error loading CSV: {e}")
        _csv_data_cache = {}

    return _csv_data_cache


def get_workbook():
    """Load the workbook with data only (calculated values)."""
    return load_workbook(EXCEL_FILE, data_only=True)


def get_workbook_formulas():
    """Load the workbook with formulas."""
    return load_workbook(EXCEL_FILE, data_only=False)


def get_task_sheet(wb):
    """Find the Task sheet or return the active sheet."""
    for sheet_name in wb.sheetnames:
        if "Task" in sheet_name:
            return wb[sheet_name]
    return wb.active


def cell_value(ws, cell):
    """
    Get cell value, preferring xlsx direct values then falling back to CSV.

    CSV fallback is needed because openpyxl cannot evaluate formulas—ssconvert
    exports calculated values to CSV which we can read.
    """
    val = ws[cell].value
    if val is not None and isinstance(val, (int, float)):
        return val
    csv_val = cell_value_csv(cell)
    if csv_val is not None:
        return csv_val
    return val if val is not None else 0


def cell_value_csv(cell):
    """Get cell value from CSV only."""
    csv_data = load_csv_data()
    return csv_data.get(cell)


def is_formula(val):
    """Check if a value is a formula string."""
    return isinstance(val, str) and val.startswith("=")


def has_formula_or_value(ws, cell):
    """Check if cell has a formula or a value from either xlsx or CSV source."""
    formula_wb = get_workbook_formulas()
    formula_ws = get_task_sheet(formula_wb)
    xlsx_val = formula_ws[cell].value
    formula_wb.close()

    if xlsx_val is not None:
        return True

    csv_val = cell_value_csv(cell)
    if csv_val is not None:
        return True

    return False


# Expected values from the Task sheet answer key
EXPECTED_EXPORTS = {
    12: [421.4, 419.5, 358.0, 432.6, 519.7, 560.6],
    13: [31.4, 30.9, 26.1, 36.6, 46.7, 41.1],
    14: [79.7, 74.5, 50.1, 80.3, 111.4, 97.8],
    15: [102.6, 92.1, 70.9, 105.6, 161.3, 127.0],
    16: [46.3, 43.6, 35.7, 46.1, 66.8, 60.9],
    17: [329.8, 303.0, 191.1, 322.1, 498.4, 418.9],
}

EXPECTED_IMPORTS = {
    19: [332.8, 333.8, 279.7, 326.0, 393.0, 439.6],
    20: [28.3, 26.4, 24.1, 28.7, 34.6, 33.1],
    21: [63.3, 56.7, 43.9, 47.0, 56.1, 63.2],
    22: [65.8, 66.8, 59.1, 61.2, 74.4, 68.5],
    23: [35.6, 32.6, 33.8, 36.9, 44.4, 45.2],
    24: [218.9, 232.1, 190.4, 240.0, 288.6, 331.6],
}

EXPECTED_NET_EXPORTS_PCT = {
    28: [11.8, 11.4, 12.3, 14.0, 13.9, 12.1],
    29: [5.1, 8.0, 4.0, 12.2, 14.8, 10.8],
    30: [11.5, 13.6, 6.6, 26.2, 33.0, 21.5],
    31: [21.8, 15.9, 9.1, 26.6, 36.9, 29.9],
    32: [13.1, 14.4, 2.7, 11.2, 20.2, 14.8],
    33: [20.2, 13.2, 0.2, 14.6, 26.7, 11.6],
}

EXPECTED_STATS = {
    35: [5.1, 8.0, 0.2, 11.2, 13.9, 10.8],
    36: [21.8, 15.9, 12.3, 26.6, 36.9, 29.9],
    37: [12.5, 13.4, 5.3, 14.3, 23.5, 13.5], 
    38: [13.9, 12.8, 5.8, 17.5, 24.2, 16.8],
    39: [11.6, 11.9, 3.0, 12.7, 16.2, 11.7], 
    40: [18.4, 14.2, 8.5, 23.3, 31.4, 19.8],
}

EXPECTED_WEIGHTED_MEAN = [15.2, 12.6, 7.4, 16.1, 22.4, 14.2]

COLUMNS = ["H", "I", "J", "K", "L", "M"]


class TestFileExists:
    """Test that the Excel file exists and is readable."""

    def test_excel_file_exists(self):
        """Verify the Excel file exists."""
        assert EXCEL_FILE.exists(), f"Excel file not found at {EXCEL_FILE}"

    def test_excel_file_readable(self):
        """Verify the Excel file can be opened."""
        wb = get_workbook()
        assert wb is not None
        wb.close()

    def test_sheet_structure(self):
        """Debug test: show available data sources."""
        wb = get_workbook()
        ws = get_task_sheet(wb)
        print(f"\nAvailable sheets: {wb.sheetnames}")
        print(f"Task sheet: {ws.title}")

        csv_file = find_task_csv()
        print(f"CSV file: {csv_file}")

        csv_data = load_csv_data()
        print(f"CSV data entries: {len(csv_data)}")

        # Print sample values from CSV
        print("\nSample CSV values:")
        for cell in ["H12", "H28", "H35", "H43"]:
            print(f"  {cell}: {csv_data.get(cell)}")

        wb.close()
        assert True


class TestStep1LookupFormulas:
    """Test Step 1: Verify lookup formulas OR values exist for data retrieval."""

    def test_exports_have_data(self):
        """Verify exports section (H12:L17) has formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(12, 18):
            for col in COLUMNS:
                cell = f"{col}{row}"
                # Check xlsx formula
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data > 0, "No data found in exports section (H12:L17)"

    def test_imports_have_data(self):
        """Verify imports section (H19:L24) has formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(19, 25):
            for col in COLUMNS:
                cell = f"{col}{row}"
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data > 0, "No data found in imports section (H19:L24)"


class TestStep1DataValues:
    """Test Step 1: Verify retrieved data values are correct."""

    def test_exports_values(self):
        """Verify exports data values match expected."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for row, expected in EXPECTED_EXPORTS.items():
            for col_idx, col in enumerate(COLUMNS):
                cell = f"{col}{row}"
                actual = cell_value(ws, cell)
                exp = expected[col_idx]
                if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                    errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "Exports value mismatches:\n" + "\n".join(errors)

    def test_imports_values(self):
        """Verify imports data values match expected."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for row, expected in EXPECTED_IMPORTS.items():
            for col_idx, col in enumerate(COLUMNS):
                cell = f"{col}{row}"
                actual = cell_value(ws, cell)
                exp = expected[col_idx]
                if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                    errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "Imports value mismatches:\n" + "\n".join(errors)


class TestStep2NetExportsCalculation:
    """Test Step 2: Net exports as % of GDP calculations."""

    def test_net_exports_pct_have_data(self):
        """Verify net exports % cells (H28:M33) have formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(28, 34):
            for col in COLUMNS:
                cell = f"{col}{row}"
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data >= 30, f"Expected at least 30 cells with data in H28:M33, found {has_data}"

    def test_net_exports_pct_values(self):
        """Verify net exports % values are calculated correctly."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for row, expected in EXPECTED_NET_EXPORTS_PCT.items():
            for col_idx, col in enumerate(COLUMNS):
                cell = f"{col}{row}"
                actual = cell_value(ws, cell)
                exp = expected[col_idx]
                if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                    errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "Net exports % mismatches:\n" + "\n".join(errors)


class TestStep2Statistics:
    """Test Step 2: Statistical calculations."""

    def test_statistics_have_data(self):
        """Verify statistics cells (H35:M40) have formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(35, 41):
            for col in COLUMNS:
                cell = f"{col}{row}"
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data >= 30, f"Expected at least 30 cells with data in H35:M40, found {has_data}"

    def test_min_values(self):
        """Verify MIN calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[35]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}35"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "MIN value mismatches:\n" + "\n".join(errors)

    def test_max_values(self):
        """Verify MAX calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[36]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}36"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "MAX value mismatches:\n" + "\n".join(errors)

    def test_median_values(self):
        """Verify MEDIAN calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[37]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}37"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "MEDIAN value mismatches:\n" + "\n".join(errors)

    def test_mean_values(self):
        """Verify AVERAGE calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[38]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}38"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "MEAN value mismatches:\n" + "\n".join(errors)

    def test_percentile_values(self):
        """Verify PERCENTILE calculation values (25th and 75th)."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected_25 = EXPECTED_STATS[39]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}39"
            actual = cell_value(ws, cell)
            exp = expected_25[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell} (25th): expected {exp}, got {actual}")

        expected_75 = EXPECTED_STATS[40]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}40"
            actual = cell_value(ws, cell)
            exp = expected_75[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell} (75th): expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "PERCENTILE value mismatches:\n" + "\n".join(errors)


class TestStep3WeightedMean:
    """Test Step 3: Weighted mean calculation using SUMPRODUCT."""

    def test_weighted_mean_has_data(self):
        """Verify weighted mean row 43 has formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for col in COLUMNS:
            cell = f"{col}43"
            if ws[cell].value is not None or csv_data.get(cell) is not None:
                has_data += 1

        wb.close()
        assert has_data >= 3, f"Expected at least 3 cells with data in row 43, found {has_data}"

    def test_weighted_mean_values_reasonable(self):
        """Verify weighted mean values are numeric and in expected range."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for _col_idx, col in enumerate(COLUMNS):
            cell = f"{col}43"
            actual = cell_value(ws, cell)

            if actual is None or actual == "" or not isinstance(actual, (int, float)):
                errors.append(f"{cell}: expected numeric value, got {actual}")
            elif actual == 0:
                errors.append(f"{cell}: value should not be zero")

        wb.close()
        assert len(errors) == 0, "Weighted mean issues:\n" + "\n".join(errors)

    def test_weighted_mean_values(self):
        """Verify weighted mean values match expected."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}43"
            actual = cell_value(ws, cell)
            exp = EXPECTED_WEIGHTED_MEAN[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "Weighted mean value mismatches:\n" + "\n".join(errors)


class TestNoExcelErrors:
    """Test that there are no Excel formula errors in the file."""

    def test_no_formula_errors(self):
        """Verify no #VALUE!, #REF!, #NAME?, etc. errors in calculated cells."""
        csv_data = load_csv_data()

        excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        error_cells = []

        # Check all relevant ranges
        ranges_to_check = [
            (12, 18),  # Exports
            (19, 25),  # Imports
            (28, 34),  # Net exports %
            (35, 41),  # Statistics
            (43, 44),  # Weighted mean
        ]

        for start_row, end_row in ranges_to_check:
            for row in range(start_row, end_row):
                for col in COLUMNS:
                    cell = f"{col}{row}"
                    val = csv_data.get(cell)
                    if val is not None and isinstance(val, str):
                        for err in excel_errors:
                            if err in str(val):
                                error_cells.append(f"{cell}: {val}")
                                break

        # Also check xlsx
        wb = get_workbook()
        ws = get_task_sheet(wb)
        for start_row, end_row in ranges_to_check:
            for row in range(start_row, end_row):
                for col in COLUMNS:
                    cell = f"{col}{row}"
                    val = ws[cell].value
                    if val is not None and isinstance(val, str):
                        for err in excel_errors:
                            if err in str(val):
                                if f"{cell}:" not in str(error_cells):
                                    error_cells.append(f"{cell}: {val}")
                                break
        wb.close()

        assert len(error_cells) == 0, "Excel errors found:\n" + "\n".join(error_cells)


EXPECTED_SHEETS = ["Task", "Data"]

EXPECTED_TASK_COLUMN_WIDTHS = {
    "A": 10.5,
    "B": 7.2,
    "C": 19.8,
    "D": 17.8,
    "E": 39.3,
    "F": 15.2,
    "G": 9.3,
    "H": 13.0,
    "I": 13.0,
    "J": 13.0,
    "K": 13.0,
    "L": 13.0,
    "M": 13.0,
}

EXPECTED_DATA_COLUMN_WIDTHS = {
    "A": 19.5,
    "B": 18.5,
    "C": 18.7,
    "D": 14.0,
    "E": 67.8,
    "F": 8.7,
    "G": 13.0,
    "H": 12.0,
    "I": 8.7,
    "J": 8.7,
    "K": 8.7,
    "L": 8.7,
    "M": 8.7,
}

WIDTH_TOLERANCE_RATIO = 0.5


class TestSheetStructure:
    """Test that no extra sheets were added to the workbook."""

    def test_no_extra_sheets_added(self):
        """Verify that only the expected sheets exist (Task and Data)."""
        wb = get_workbook()
        actual_sheets = wb.sheetnames
        wb.close()

        extra_sheets = [s for s in actual_sheets if s not in EXPECTED_SHEETS]

        assert len(extra_sheets) == 0, (
            f"Extra sheets were added to the workbook: {extra_sheets}. "
            f"Only {EXPECTED_SHEETS} should exist. "
            f"Found sheets: {actual_sheets}"
        )

    def test_required_sheets_exist(self):
        """Verify that the required sheets (Task and Data) still exist."""
        wb = get_workbook()
        actual_sheets = wb.sheetnames
        wb.close()

        missing_sheets = [s for s in EXPECTED_SHEETS if s not in actual_sheets]

        assert len(missing_sheets) == 0, (
            f"Required sheets are missing: {missing_sheets}. " f"Expected sheets: {EXPECTED_SHEETS}. " f"Found sheets: {actual_sheets}"
        )


class TestFormattingPreserved:
    """Test that formatting (column widths) was not drastically changed."""

    def test_task_sheet_column_widths(self):
        """Verify Task sheet column widths haven't been drastically changed."""
        wb = get_workbook()

        task_ws = None
        for sheet_name in wb.sheetnames:
            if "Task" in sheet_name or "task" in sheet_name.lower():
                task_ws = wb[sheet_name]
                break

        if task_ws is None:
            wb.close()
            pytest.skip("Task sheet not found")

        # ssconvert normalizes widths to ~1.7 as a conversion artifact—skip if detected
        widths = [task_ws.column_dimensions[col].width or 8.43 for col in EXPECTED_TASK_COLUMN_WIDTHS]
        if any(w < 3 for w in widths):
            wb.close()
            pytest.skip("Column widths appear to be normalized by ssconvert (conversion artifact)")

        errors = []
        for col, expected_width in EXPECTED_TASK_COLUMN_WIDTHS.items():
            actual_width = task_ws.column_dimensions[col].width
            # Handle None (default width is typically ~8.43)
            if actual_width is None:
                actual_width = 8.43

            # Calculate tolerance bounds
            min_width = expected_width * (1 - WIDTH_TOLERANCE_RATIO)
            max_width = expected_width * (1 + WIDTH_TOLERANCE_RATIO)

            if actual_width < min_width or actual_width > max_width:
                errors.append(
                    f"Column {col}: expected width ~{expected_width:.1f}, "
                    f"got {actual_width:.1f} (outside {min_width:.1f}-{max_width:.1f} range)"
                )

        wb.close()
        assert len(errors) == 0, (
            "Task sheet column widths were drastically changed:\n"
            + "\n".join(errors)
            + "\nFormatting should be preserved for human readability."
        )

    def test_data_sheet_column_widths(self):
        """Verify Data sheet column widths haven't been drastically changed."""
        wb = get_workbook()

        data_ws = None
        for sheet_name in wb.sheetnames:
            if sheet_name == "Data":
                data_ws = wb[sheet_name]
                break

        if data_ws is None:
            wb.close()
            pytest.skip("Data sheet not found")

        # ssconvert normalizes widths to ~1.7 as a conversion artifact—skip if detected
        widths = [data_ws.column_dimensions[col].width or 8.43 for col in EXPECTED_DATA_COLUMN_WIDTHS]
        if any(w < 3 for w in widths):
            wb.close()
            pytest.skip("Column widths appear to be normalized by ssconvert (conversion artifact)")

        errors = []
        for col, expected_width in EXPECTED_DATA_COLUMN_WIDTHS.items():
            actual_width = data_ws.column_dimensions[col].width
            # Handle None (default width is typically ~8.43)
            if actual_width is None:
                actual_width = 8.43

            # Calculate tolerance bounds
            min_width = expected_width * (1 - WIDTH_TOLERANCE_RATIO)
            max_width = expected_width * (1 + WIDTH_TOLERANCE_RATIO)

            if actual_width < min_width or actual_width > max_width:
                errors.append(
                    f"Column {col}: expected width ~{expected_width:.1f}, "
                    f"got {actual_width:.1f} (outside {min_width:.1f}-{max_width:.1f} range)"
                )

        wb.close()
        assert len(errors) == 0, (
            "Data sheet column widths were drastically changed:\n"
            + "\n".join(errors)
            + "\nFormatting should be preserved for human readability."
        )


class TestNoMacros:
    """Test that no VBA macros were introduced in the Excel file."""

    def test_no_vba_macros(self):
        """Verify the Excel file does not contain VBA macro code."""
        import zipfile

        with zipfile.ZipFile(EXCEL_FILE, "r") as zf:
            vba_files = [n for n in zf.namelist() if "vbaProject" in n or n.endswith(".bin")]

        assert len(vba_files) == 0, (
            f"The Excel file contains VBA macro code: {vba_files}. " "Macros are not allowed - please use only Excel formulas."
        )

    def test_file_extension_xlsx(self):
        """Verify the file is .xlsx (not .xlsm which supports macros)."""
        assert EXCEL_FILE.suffix.lower() == ".xlsx", (
            f"Expected .xlsx file extension, got {EXCEL_FILE.suffix}. " "Macro-enabled formats (.xlsm) are not allowed."
        )
